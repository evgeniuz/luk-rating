import csv
import openpyxl
import openskill.models

mu = 25
sigma = mu / 3.0

beta = sigma * 4.5
tau = sigma / 100.0


def gamma(*args, **kwargs):
    return 1.0


def ordinal(rating):
    return rating.ordinal(z=2.0, alpha=5.0, target=1000.0)


model = openskill.models.BradleyTerryFull(
    mu=mu, sigma=sigma, beta=beta, tau=tau, gamma=gamma
)

TEAM_MAPPING = {
    "Цинiчнi бандери": ["Цинічні бандери"],
    "Beit Grand - Аборигени": ["Beit Grand - Аборигены", "Аборигени", "Аборигени+"],
    "Автопілот": ["Автопилот"],
    "Мінус один": ["Минус один"],
    "Соснова шишечка": ["Сосновая шишечка", "Соснова Шишечка"],
    "Прозорі Зумери": ["Прозорі Зззумери"],
    "Капітан Джек": ["Капитан Джек", "Captain Jack"],
    "Дракони Ймовірності": ["Drakony Jmovirnosti"],
    "Забобонні бонобо": ["Забобонні Бонобо"],
    "Kéisécker": ["Keisecker"],
    "seꑭes": ["senes"],
    "Манія Величі": ["Мания Величия"],
    "Злобні Урук-хай": ["Злобні урукхаї", "Злобные Урук-хай"],
    "Харківська весільна слононіжка": ["Харківська Весільна Слононіжка", "Слононіжка"],
    "Яка вам різниця?": ["Яка Вам Різниця", "Яка Вам Різниця?"],
    "Гря хм гагага": ["гря хм гагага"],
    "В'язні міста ІФ": ["В'язні міста Іф"],
    "Сливки - Форс-Мажор": ["СФМ"],
    "Балабеники": ["28 тисяч вивірок"],
    "БЗ": ["Б.З."],
    "Команда №2": ["Команда номер 2"],
    "Who knew?": ["Who Knew?"],
    "Де Лореан?": ["Ми з Миколаєва"],
    "Західний полюс": ["Захидный полюс", "Західний Полюс"],
    "Генератор": ["Генератор случайных слов"],
    "Пунктуальність": ["Пунктуальность"],
    "Коментарі Асгарда": ["Комментарии Асгарда"],
    "Каїсса - Дикий Сад": ["Каисса - Дикий сад"],
    "Черкаське FIDO": ["Черкасское FIDO"],
    "Галера надвечір": ["Галена надвечір"],
    "Серьога Хруст і Гідрогарік": ["Серёга Хруст и ГидрогарикTwo "],
    "Алкобоги": ["Алокбоги"],
    "Lorem City": ["Lorem Ipsum"],
}

EXCLUDED_TEAMS = [
    "Пупи та Лупи",
    "Склад фʼючерсної свинини",
    "Збірна імені Патріка Стара",
    "Вскуйовджені Хохулі",
    "GOL",
    "Лама",
    "Неможна в ілюмінаторі",
    "Тжвжик",
    "This is Sparжааа",
    "Sich",
    "Soles",
    "BBQ test 300",
    "Чайка на ім'я мартин Іден",
    "КоньякСпорт",
    "Playboy Україна",
    "В пошуках Дорі",
    "Нет оторвы",
    "ЗМЗ 2024",
    "Тореадори з Країни Сонячних Зайчиків",
    "Бар Міцва",
    "Rule34",
    "ДНК",
    "Андрюша",
]


def game1():
    workbook = openpyxl.load_workbook("raw/1_етап_ЧУ-О3.xlsx")
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=3, values_only=True):
        name = row[1]

        score = 0
        score += sum((i for i in row[3 : 3 + 12] if i is not None))
        score += sum((i for i in row[16 : 16 + 12] if i is not None))
        score += sum((i for i in row[29 : 29 + 12] if i is not None))

        yield name, score


def game2():
    workbook = openpyxl.load_workbook("raw/СинЛУК - Зима - результат.xlsx")
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        score = row[2]

        yield name, score


def game3():
    workbook = openpyxl.load_workbook("raw/Результати Зимовий Мажор.xlsx")
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[1]
        score = row[2]

        if name is None:
            continue

        yield name, score


def game4():
    with open("raw/tovstolobyk.csv", "r", encoding="utf-8") as file:
        reader = csv.reader(file, delimiter=",")
        next(reader)

        for row in reader:
            name = row[0]
            score = int(row[1])

            yield name, score


def game5():
    workbook = openpyxl.load_workbook("raw/Результати СинЛУК - Весна.xlsx")
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        score = row[3]

        yield name, score


def game6():
    workbook = openpyxl.load_workbook("raw/Результати мінор.xlsx")
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        tiebreak = row[7]
        score = row[8]

        if tiebreak is not None:
            score += tiebreak

        yield name, score


def game7():
    workbook = openpyxl.load_workbook("raw/Результати Simply Green з Mamihlapinatapai.xlsx")
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        score = row[2]

        yield name, score


GAMES = {
    "Гра 1": {"loader": game1, "weight": 1},
    "Гра 2": {"loader": game2, "weight": 1},
    "Гра 3": {"loader": game3, "weight": 2},
    "Гра 4": {"loader": game4, "weight": 1},
    "Гра 5": {"loader": game5, "weight": 1},
    "Гра 6": {"loader": game6, "weight": 2},
    "Гра 7": {"loader": game7, "weight": 1},
}


def preprocess(results):
    variant_to_canonical = {}
    for canonical, variants in TEAM_MAPPING.items():
        for variant in variants:
            variant_to_canonical[variant] = canonical

    for team, score in results:
        trimmed = team.strip()

        canonical_name = variant_to_canonical.get(trimmed, trimmed)

        if canonical_name in EXCLUDED_TEAMS:
            continue

        yield canonical_name, score


def calculate_team_ranks(results):
    sorted_results = sorted(results, key=lambda x: x[1], reverse=True)
    teams = []
    scores = []
    ranks = []

    last_score = None
    last_rank = 0

    for i, (team, score) in enumerate(sorted_results, start=1):
        teams.append(team)
        scores.append(score)

        if last_score is not None and score == last_score:
            ranks.append(last_rank)
            continue

        ranks.append(i)
        last_rank = i
        last_score = score

    return teams, scores, ranks


def format_rating(rating):
    if rating is None:
        return "", "", ""
    return f"{rating.mu:.4f}", f"{rating.sigma:.4f}", f"{ordinal(rating):.2f}"


def rating():
    history = {}
    ratings = {}

    for game_id, game in GAMES.items():
        results = preprocess(game["loader"]())

        teams, scores, ranks = calculate_team_ranks(results)

        for team in teams:
            if team not in ratings:
                ratings[team] = model.rating()

        team_ratings = [[ratings[team]] for team in teams]

        for _ in range(game["weight"]):
            team_ratings = model.rate(team_ratings, ranks=ranks)

        history[game_id] = {}

        for team, score, rank, rating in zip(teams, scores, ranks, team_ratings):
            ratings[team] = rating[0]

            history[game_id][team] = {
                "score": score,
                "rank": rank,
                "rating": rating[0],
            }

    return ratings, history


def export(ratings, history, filename="history.csv"):
    sorted_teams = [
        team
        for team, _ in sorted(
            ratings.items(), key=lambda item: ordinal(item[1]), reverse=True
        )
    ]

    header = ["Команда"]
    header.append("Поточний рейтинг")
    for game_id, game in GAMES.items():
        multiplier = f" (×{game['weight']})" if game["weight"] > 1 else ""

        header.extend(
            [
                f"{game_id} Результат",
                f"{game_id} Місце",
                f"{game_id} μ{multiplier}",
                f"{game_id} σ{multiplier}",
                f"{game_id} Рейтинг{multiplier}",
            ]
        )

    with open(filename, "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(header)
        for team in sorted_teams:
            final = ratings.get(team, None)
            _, _, final_formatted_rating = format_rating(final)

            row = [team, final_formatted_rating]
            for game in GAMES.keys():
                details = history.get(game, {}).get(team, {})

                score = details.get("score", "")
                rank = details.get("rank", "")
                rating = details.get("rating", None)

                mu, sigma, formatted_rating = format_rating(rating)

                row.extend([score, rank, mu, sigma, formatted_rating])
            writer.writerow(row)


if __name__ == "__main__":
    ratings, history = rating()
    export(ratings, history)
